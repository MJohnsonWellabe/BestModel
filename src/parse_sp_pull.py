"""Parse a raw S&P 'Life/Fraternal Financial Highlights' export into a tidy record.

The export is the analyst's backbone pull (one .xlsx per carrier in data/raw/). Its row
labels are stable across the sample carriers but may drift for atypical entities (P&C
siblings, pure-health), so rows are matched by a normalized **alias dictionary** rather
than fixed positions, and any label we can't place is reported as `unmatched` for the
analyst to map.

Returns a RawPull with:
  - name / entity_key  (identity, cross-checked against the roster)
  - years              (the full calendar years present, e.g. [2022, 2023, 2024, 2025])
  - series[target]     = {year: value}   for every matched metric
  - unmatched          = [raw labels we could not place]   (drift surfacing)

Downstream, features.py turns series[] into the normalized schema (latest-year scalars,
5-yr stats, and the derived metrics). Nothing here invents values: absent rows are simply
absent from series[], and the build's missingness report makes that visible.
"""
from __future__ import annotations

import datetime as dt
import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl


def normalize(label) -> str:
    """Lowercase, strip whitespace/punctuation so label variants compare equal."""
    return re.sub(r"[^a-z0-9]", "", str(label).lower())


# target_key -> list of acceptable source row labels (normalized at match time).
# Order within a list = preference; first match wins. Add aliases here when the
# build report flags a drifted label for some carrier.
ALIASES: dict[str, list[str]] = {
    # --- balance sheet, $000 (latest-year scalar) ---
    # Life/Fraternal uses "and"; Health uses "&". P&C shares the Life labels.
    "total_invested_assets_usd": ["Total Cash and Investments", "Total Cash & Investments"],
    "total_assets_usd": ["Total Assets"],
    "affiliated_investments_usd": ["Affiliated Investments (incl above)"],
    "total_reserves_usd": ["Total Policy Reserves"],
    "avr_usd": ["Asset Valuation Reserve"],
    "surplus_notes_usd": ["Surplus Notes"],
    "capital_surplus_usd": ["Capital and Surplus", "Capital & Surplus"],
    # --- RBC, $000 + ratio ---
    "tac_usd": ["RBC - Total Adjusted Capital", "Total Adjusted Capital"],
    "acl_rbc_usd": ["ACL Risk Based Capital", "Authorized Control Level Risk Based Capital"],
    "rbc_acl_ratio": ["Risk Based Capital Ratio(TAC/ACL RBC)",
                      "Risk Based Capital Ratio (TAC/ACL RBC) (%)", "Risk Based Capital Ratio"],
    # health template combined ratio (used when life benefit/expense/commission rows are absent)
    "combined_ratio_direct": ["Combined Ratio"],
    # --- asset quality / mix (%, latest-year) ---
    "bonds_pct": ["Unaff. Bonds / Unaff. Investments"],
    "common_stock_pct": ["Unaff. Common Stocks / Unaff. Investments"],
    "preferred_stock_pct": ["Unaff. Preferred Stocks / Unaff. Investments"],
    "affiliated_pct": ["Affiliated Investments / Total Investments"],
    "below_ig_pct": ["Bonds Rated 3-6 / Total Bonds"],
    "bond_avg_quality": ["Bond Average Asset Quality (1-6) (#)", "Bond Average Asset Quality"],
    # --- income statement, $000 (series; latest used as current) ---
    "total_premium": ["Premiums, Consideration & Deposits"],
    "net_investment_income": ["Net Investment Income"],
    "realized_gains": ["Net Realized Capital Gains (Losses)", "Net Realized Capital Gains"],
    "net_income": ["Net Income"],
    "pretax_op_income": ["Pre-tax Operating Income"],
    "benefits": ["Benefits"],
    # --- operating ratios (%, full series) ---
    "roe": ["Return on Average Equity"],
    "roa": ["Return on Average Assets"],
    "benefit_ratio": ["Benefit Ratio (Premiums)"],
    "expense_ratio": ["Expense Ratio (Premiums)"],
    "commission_ratio": ["Commission Ratio"],
    "premium_growth": ["Growth Rate - Premium & Annuity Cnsdrtns"],
    "net_premiums_to_cs": ["Net Premiums Written / C&S"],
    # --- segment mix (%, latest) ---
    "seg_individual_life": ["Individual Life"],
    "seg_group_life": ["Group Life"],
    "seg_individual_annuity": ["Individual Annuities"],
    "seg_group_annuity": ["Group Annuities"],
    "seg_ah": ["Accident & Health"],
    # --- reinsurance ceded reserve credits ($000) — proxy for recoverable leverage ---
    "reins_credit_life_ga": ["Life & Annuities General Accounts"],
    "reins_credit_ah": ["Accident & Health (incl Unearned Prem)"],
}

# Build a reverse lookup: normalized source label -> target_key.
_NORM_TO_TARGET: dict[str, str] = {}
for _t, _labels in ALIASES.items():
    for _l in _labels:
        _NORM_TO_TARGET.setdefault(normalize(_l), _t)


@dataclass
class RawPull:
    name: str
    entity_key: str | None
    source_file: str
    years: list[int] = field(default_factory=list)
    series: dict[str, dict[int, float]] = field(default_factory=dict)
    unmatched: list[str] = field(default_factory=list)


def _is_highlights(ws) -> bool:
    """Signature check: the export self-identifies in its first rows."""
    head = " ".join(
        normalize(c)
        for row in ws.iter_rows(min_row=1, max_row=12, values_only=True)
        for c in row
        if c is not None
    )
    return "financialhighlights" in head or "periodended" in head


def _year_columns(ws) -> dict[int, int]:
    """Map worksheet column index -> calendar year, using the 'Period Ended' row.

    Only full calendar years (month == December) are kept; YTD columns are dropped.
    Falls back to scanning the header band for 4-digit years if no date row is found.
    """
    for row in ws.iter_rows(values_only=False):
        first = row[0].value
        if first is not None and normalize(first) == normalize("Period Ended"):
            cols: dict[int, int] = {}
            for cell in row[1:]:
                v = cell.value
                if isinstance(v, dt.datetime) and v.month == 12:
                    cols[cell.column] = v.year
            if cols:
                return cols
    # fallback: header strings like "2024 Y"
    cols = {}
    for row in ws.iter_rows(min_row=1, max_row=12, values_only=False):
        for cell in row[1:]:
            m = re.search(r"(19|20)\d{2}", str(cell.value or ""))
            if m and "ytd" not in str(cell.value).lower():
                cols[cell.column] = int(m.group())
    return cols


def parse_highlights(path: str | Path) -> RawPull:
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = next((w for w in wb.worksheets if _is_highlights(w)), wb.active)

    # identity: row with the company name is the first non-empty A-cell containing " | " or a name
    name, entity_key = path.stem, None
    for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
        a = row[0]
        if a and "|" in str(a):
            name = str(a).split("|")[0].strip()
            break
    for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
        a = str(row[0] or "")
        m = re.search(r"STAT ENTITY KEY:\s*(\S+)", a)
        if m:
            entity_key = m.group(1)
            break

    ycols = _year_columns(ws)
    years = sorted(set(ycols.values()))
    pull = RawPull(name=name, entity_key=entity_key, source_file=path.name, years=years)

    seen_targets: set[str] = set()
    for row in ws.iter_rows(values_only=False):
        label = row[0].value
        if label is None or str(label).strip() == "":
            continue
        target = _NORM_TO_TARGET.get(normalize(label))
        if target is None:
            # only report data-ish rows (skip section headers / notes)
            if any(
                isinstance(c.value, (int, float)) for c in row[1:6]
            ) and target is None:
                pull.unmatched.append(str(label).strip())
            continue
        if target in seen_targets:
            continue  # first match wins (handles duplicate sub-rows)
        seen_targets.add(target)
        by_year: dict[int, float] = {}
        for cell in row[1:]:
            yr = ycols.get(cell.column)
            if yr is None:
                continue
            v = cell.value
            if isinstance(v, (int, float)):
                by_year[yr] = float(v)
        if by_year:
            pull.series[target] = by_year

    return pull


if __name__ == "__main__":  # quick manual check
    import sys

    for p in sys.argv[1:] or sorted(Path("data/raw").glob("*.xlsx")):
        r = parse_highlights(p)
        print(f"\n=== {r.name}  (entity {r.entity_key}, years {r.years}) ===")
        print(f"  matched {len(r.series)} metrics; unmatched {len(r.unmatched)}")
        for k in ("tac_usd", "acl_rbc_usd", "rbc_acl_ratio", "capital_surplus_usd",
                  "below_ig_pct", "roe"):
            if k in r.series:
                print(f"    {k:20} {r.series[k]}")
